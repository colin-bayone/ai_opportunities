#!/usr/bin/env python3
"""
Scratchpad for McGrath RFP Excel processing experiments.
Run from: mcgrath/rfp_docs/
"""

import csv
from openpyxl import load_workbook
from pathlib import Path

FILE = Path(__file__).parent / "McGrath RentCorp_Managed Services Provider RFP_02172026.xlsx"
OUTPUT_DIR = Path(__file__).parent / "extracted"


def scan_sheets():
    """Quick scan of all sheets - should be fast."""
    print(f"Loading: {FILE.name}")
    print(f"Size: {FILE.stat().st_size / 1024 / 1024:.2f} MB\n")

    wb = load_workbook(FILE, read_only=True, data_only=True)

    print(f"{'Sheet Name':<30} {'Max Row':<10} {'Max Col':<10}")
    print("-" * 50)

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"{name:<30} {ws.max_row:<10} {ws.max_column:<10}")

    wb.close()
    print("\nDone.")


def extract_sheet_to_csv(sheet_name: str, max_cols: int = 20):
    """
    Extract a single sheet to CSV.
    max_cols limits columns to avoid the 16384 issue.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / f"{sheet_name.lower().replace(' ', '_')}.csv"

    print(f"Extracting '{sheet_name}' to {output_file}...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        for row in ws.iter_rows(max_col=max_cols):
            row_values = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_values.append("")
                else:
                    # Clean up the value
                    row_values.append(str(val).strip())

            # Skip completely empty rows
            if any(v for v in row_values):
                writer.writerow(row_values)
                rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def extract_specific_requirements_clean():
    """
    Extract Specific Requirements with:
    - Skip header rows (start at row 11, headers in row 10)
    - Propagate Solution column down
    - Clean True/False to Yes/No
    - Replace newlines in text with spaces
    - Remove duplicate Solution column
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "specific_requirements_clean.csv"

    print(f"Extracting 'Specific Requirements' (cleaned)...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["Specific Requirements"]

    # Column mapping (1-indexed in Excel)
    # B=Solution, C=Category, D=Specific Requirements,
    # E=MSP Owned, F=Shared, G=MSP Augmented, H=MGRC Owned,
    # I=Decline to Bid, J=Bid As Specified, K=Bid As Modified, L=MSP Comments

    headers = [
        "Solution", "Category", "Specific Requirements",
        "MSP Owned", "Shared", "MSP Augmented", "MGRC Owned",
        "Decline to Bid", "Bid As Specified", "Bid As Modified", "MSP Comments"
    ]

    rows_written = 0
    current_solution = ""

    def clean_text(val):
        """Clean text value - replace newlines, strip whitespace."""
        if val is None:
            return ""
        text = str(val).strip()
        # Replace newlines with spaces
        text = text.replace('\n', ' ').replace('\r', ' ')
        # Collapse multiple spaces
        while '  ' in text:
            text = text.replace('  ', ' ')
        return text

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        # Start at row 5 (after headers in row 4)
        for row in ws.iter_rows(min_row=5, min_col=2, max_col=12):
            values = [cell.value for cell in row]

            # Get solution (column B = index 0)
            solution = values[0]
            if solution:
                current_solution = clean_text(solution)

            # Get category (column C = index 1)
            category = values[1]
            if not category:
                continue  # Skip rows without category

            # Build clean row
            clean_row = [
                current_solution,
                clean_text(values[1]),  # Category
                clean_text(values[2]),  # Specific Requirements
            ]

            # Convert True/False to Yes/No for checkbox columns
            for i in range(3, 10):
                val = values[i]
                if val is True:
                    clean_row.append("Yes")
                elif val is False:
                    clean_row.append("No")
                else:
                    clean_row.append("")

            # MSP Comments
            clean_row.append(clean_text(values[10]))

            writer.writerow(clean_row)
            rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")

    # Print unique solutions found
    print("\nUnique Solutions found:")
    solutions = set()
    with open(output_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Solution']:
                solutions.add(row['Solution'])
    for s in sorted(solutions):
        print(f"  - {s}")

    return output_file


def clean_text(val):
    """Clean text value - replace newlines, strip whitespace."""
    if val is None:
        return ""
    text = str(val).strip()
    text = text.replace('\n', ' ').replace('\r', ' ')
    while '  ' in text:
        text = text.replace('  ', ' ')
    return text


def extract_general_requirements_clean():
    """
    Extract General Requirements tab with:
    - Skip title row (row 1) and empty row (row 2)
    - Column B = Requirement Number, C = Requirement Text, D = MSP Response
    - Identify section headers (whole numbers) vs requirements (decimals)
    - Clean newlines from text
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "general_requirements_clean.csv"

    print(f"Extracting 'General Requirements' (cleaned)...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["General Requirements"]

    headers = ["Section", "Number", "Requirement", "Type", "MSP Response"]

    rows_written = 0
    current_section = ""

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        # Start at row 3 (after title in row 1, empty row 2)
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=4):
            values = [cell.value for cell in row]

            # Column B = Number (index 0)
            number = values[0]
            if number is None:
                continue  # Skip empty rows

            number_str = str(number).strip()
            requirement = clean_text(values[1])
            response = clean_text(values[2])

            # Determine if section header or requirement
            # Section headers are whole numbers (1, 2, 3...)
            # Requirements have decimals (1.1, 1.2, 2.1...)
            if '.' in number_str:
                row_type = "Requirement"
                section = current_section
            else:
                row_type = "Section Header"
                current_section = requirement  # The requirement text IS the section name
                section = ""

            clean_row = [section, number_str, requirement, row_type, response]
            writer.writerow(clean_row)
            rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")

    # Summary stats
    print("\nSummary:")
    sections = 0
    requirements = 0
    with open(output_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Type'] == 'Section Header':
                sections += 1
            else:
                requirements += 1
    print(f"  Sections: {sections}")
    print(f"  Requirements: {requirements}")

    return output_file


def extract_sla_clean():
    """
    Extract SLA tab - matrix format with priority levels as columns.
    Row 3 = headers, Rows 4-11 = data
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "sla_clean.csv"

    print(f"Extracting 'SLA' (cleaned)...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["SLA"]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Extract rows 3-11 (header + data), columns B-F
        for row in ws.iter_rows(min_row=3, max_row=11, min_col=2, max_col=6):
            values = [clean_text(cell.value) for cell in row]

            # Skip empty rows
            if not any(values):
                continue

            writer.writerow(values)
            rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def extract_kpi_clean():
    """
    Extract KPI tab - simple 2-column table.
    Row 3 = headers, Rows 4-11 = data
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "kpi_clean.csv"

    print(f"Extracting 'KPI' (cleaned)...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["KPI"]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Extract rows 3-11 (header + data), columns B-C
        for row in ws.iter_rows(min_row=3, max_row=11, min_col=2, max_col=3):
            values = [clean_text(cell.value) for cell in row]

            # Skip empty rows
            if not any(values):
                continue

            writer.writerow(values)
            rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def debug_first_rows(sheet_name: str = "Specific Requirements", max_cols: int = 8):
    """Debug: show first 15 rows of a sheet to find header row."""
    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb[sheet_name]

    print(f"First 15 rows of '{sheet_name}' (cols A-{chr(64+max_cols)}):\n")
    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=max_cols), start=1):
        values = [str(cell.value)[:40] if cell.value else "---" for cell in row]
        print(f"Row {row_num:2d}: {values}")

    wb.close()


def extract_overview_clean():
    """
    Extract RFP MGRC Overview - prose/narrative content.
    Single column of text with section headers.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "rfp_overview.csv"

    print(f"Extracting 'RFP MGRC Overview'...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["RFP MGRC Overview"]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Row", "Content"])

        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=80, min_col=2, max_col=2), start=1):
            val = row[0].value
            if val:
                text = clean_text(val)
                if text:
                    writer.writerow([row_num, text])
                    rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def extract_pricing_template():
    """
    Extract Pricing Here tab - template structure with headers.
    Captures the pricing grid structure for reference.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "pricing_template.csv"

    print(f"Extracting 'Pricing Here' (template structure)...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["Pricing Here"]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=60, min_col=2, max_col=8), start=1):
            values = [clean_text(cell.value) for cell in row]
            if any(values):
                writer.writerow([row_num] + values)
                rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def extract_next_steps_clean():
    """
    Extract Next Steps tab - prose/narrative content.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "next_steps.csv"

    print(f"Extracting 'Next Steps'...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["Next Steps"]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(["Row", "Content"])

        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=55, min_col=2, max_col=3), start=1):
            values = [clean_text(cell.value) for cell in row]
            combined = " | ".join([v for v in values if v])
            if combined:
                writer.writerow([row_num, combined])
                rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def extract_experience_methodology():
    """
    Extract Experience & Methodology tab - questionnaire format.
    Same structure as General Requirements.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "experience_methodology.csv"

    print(f"Extracting 'Experience & Methodology'...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["Experience & Methodology"]

    headers = ["Section", "Number", "Question", "Type", "Response"]

    rows_written = 0
    current_section = ""

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for row in ws.iter_rows(min_row=3, min_col=2, max_col=4):
            values = [cell.value for cell in row]

            number = values[0]
            if number is None:
                continue

            number_str = str(number).strip()
            question = clean_text(values[1])
            response = clean_text(values[2])

            if '.' in number_str:
                row_type = "Question"
                section = current_section
            else:
                row_type = "Section Header"
                current_section = question
                section = ""

            writer.writerow([section, number_str, question, row_type, response])
            rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


def extract_corporate_response():
    """
    Extract Corporate Response Form - questionnaire with multiple columns.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file = OUTPUT_DIR / "corporate_response_form.csv"

    print(f"Extracting 'Corporate Response Form'...")

    wb = load_workbook(FILE, read_only=True, data_only=True)
    ws = wb["Corporate Response Form"]

    rows_written = 0
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Start at row 3 (headers), columns B-F
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=6):
            values = [clean_text(cell.value) for cell in row]

            # Skip empty rows
            if not any(values):
                continue

            writer.writerow(values)
            rows_written += 1

    wb.close()
    print(f"Wrote {rows_written} rows to {output_file}")
    return output_file


if __name__ == "__main__":
    extract_experience_methodology()
    extract_corporate_response()
