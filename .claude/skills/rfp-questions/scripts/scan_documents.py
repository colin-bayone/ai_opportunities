#!/usr/bin/env python3
"""
Document Scanner for RFP Questions Skill

Inventories source documents in an RFP session folder.
Reports file types, sizes, and page counts (where detectable).
Never reads content - just metadata for planning.

Usage:
    python3 scan_documents.py <session_folder>
    python3 scan_documents.py .rfp-questions/acme-corp

Output: JSON to stdout with document inventory
"""

import json
import os
import sys
from pathlib import Path
from typing import Any


def get_pdf_page_count(file_path: Path) -> int | None:
    """Attempt to get PDF page count without full parsing."""
    try:
        # Simple regex-based count (works for most PDFs)
        content = file_path.read_bytes()
        # Count /Type /Page occurrences (rough estimate)
        count = content.count(b"/Type/Page") + content.count(b"/Type /Page")
        return count if count > 0 else None
    except Exception:
        return None


def get_excel_sheet_info(file_path: Path) -> dict | None:
    """Get Excel sheet names and row counts without full parsing."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Get dimensions if available
            if ws.dimensions:
                sheets[sheet_name] = {"dimensions": ws.dimensions}
            else:
                sheets[sheet_name] = {"dimensions": "unknown"}
        wb.close()
        return sheets
    except ImportError:
        return {"error": "openpyxl not installed"}
    except Exception as e:
        return {"error": str(e)}


def scan_file(file_path: Path) -> dict[str, Any]:
    """Scan a single file for metadata."""
    stat = file_path.stat()
    suffix = file_path.suffix.lower()

    info = {
        "name": file_path.name,
        "path": str(file_path),
        "size_bytes": stat.st_size,
        "size_human": format_size(stat.st_size),
        "type": suffix.lstrip(".") or "unknown",
    }

    # Type-specific metadata
    if suffix == ".pdf":
        info["category"] = "document"
        page_count = get_pdf_page_count(file_path)
        if page_count:
            info["pages"] = page_count
            info["processing_note"] = "Use Gemini native PDF processing"

    elif suffix in (".xlsx", ".xls"):
        info["category"] = "spreadsheet"
        sheet_info = get_excel_sheet_info(file_path)
        if sheet_info:
            info["sheets"] = sheet_info
        info["processing_note"] = "Use openpyxl + Gemini for intelligent extraction"

    elif suffix in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        info["category"] = "image"
        info["processing_note"] = "Use Gemini vision for analysis"

    elif suffix in (".csv", ".tsv"):
        info["category"] = "data"
        info["processing_note"] = "Direct parsing possible"

    elif suffix in (".doc", ".docx"):
        info["category"] = "document"
        info["processing_note"] = "Convert to PDF or use python-docx"

    elif suffix in (".txt", ".md"):
        info["category"] = "text"
        info["processing_note"] = "Direct reading"

    else:
        info["category"] = "other"
        info["processing_note"] = "Manual review recommended"

    return info


def format_size(size_bytes: int) -> str:
    """Format bytes as human-readable size."""
    for unit in ["B", "KB", "MB", "GB"]:
        if size_bytes < 1024:
            return f"{size_bytes:.1f} {unit}"
        size_bytes /= 1024
    return f"{size_bytes:.1f} TB"


def scan_folder(folder: Path) -> dict[str, Any]:
    """Scan a folder for documents."""
    source_folder = folder / "source"

    if not source_folder.exists():
        return {
            "error": f"Source folder not found: {source_folder}",
            "suggestion": "Create 'source/' subfolder and add documents"
        }

    files = []
    total_size = 0
    categories = {}

    for file_path in sorted(source_folder.iterdir()):
        if file_path.is_file() and not file_path.name.startswith("."):
            info = scan_file(file_path)
            files.append(info)
            total_size += info["size_bytes"]

            cat = info.get("category", "other")
            categories[cat] = categories.get(cat, 0) + 1

    return {
        "session_folder": str(folder),
        "source_folder": str(source_folder),
        "total_files": len(files),
        "total_size": format_size(total_size),
        "total_size_bytes": total_size,
        "categories": categories,
        "files": files,
        "recommendations": generate_recommendations(files, total_size)
    }


def generate_recommendations(files: list, total_size: int) -> list[str]:
    """Generate processing recommendations based on scan results."""
    recs = []

    # Size warnings
    if total_size > 50 * 1024 * 1024:  # > 50MB
        recs.append("Large document set. Consider processing in batches.")

    # Count types
    pdfs = [f for f in files if f.get("type") == "pdf"]
    excels = [f for f in files if f.get("type") in ("xlsx", "xls")]
    images = [f for f in files if f.get("category") == "image"]

    if pdfs:
        total_pages = sum(f.get("pages", 0) for f in pdfs)
        if total_pages > 100:
            recs.append(f"~{total_pages} PDF pages. Use selective page processing.")
        recs.append(f"{len(pdfs)} PDF(s) - Gemini native processing recommended.")

    if excels:
        recs.append(f"{len(excels)} Excel file(s) - Review sheets before extraction.")

    if images:
        recs.append(f"{len(images)} image(s) - Gemini vision for text extraction.")

    if not recs:
        recs.append("Document set looks manageable for processing.")

    return recs


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scan_documents.py <session_folder>", file=sys.stderr)
        print("Example: python3 scan_documents.py .rfp-questions/acme-corp", file=sys.stderr)
        sys.exit(1)

    folder = Path(sys.argv[1])

    if not folder.exists():
        print(json.dumps({"error": f"Folder not found: {folder}"}))
        sys.exit(1)

    result = scan_folder(folder)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
