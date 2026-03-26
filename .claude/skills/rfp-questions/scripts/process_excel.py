#!/usr/bin/env python3
"""
Excel Processor for RFP Questions Skill

Uses openpyxl for structure analysis and Gemini for intelligent CSV extraction.
Produces sheet-by-sheet breakdown with sensible CSV conversions.

Usage:
    python3 process_excel.py <file_path> --output-dir <dir> [--sheets Sheet1,Sheet2]

Environment:
    GEMINI_API_KEY - Required for intelligent extraction guidance (can be in .env file)

Output:
    Creates files in output directory:
    - index.md (workbook overview with relationships)
    - sheet_<name>.md (markdown representation of each sheet)
    - sheet_<name>.csv (CSV export, possibly multiple if complex)
    - relationships.json (data relationships discovered)
"""

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any


def load_env():
    """Load environment variables from .env file if available."""
    try:
        from dotenv import load_dotenv
        # Look for .env in current dir and parent dirs
        env_path = Path.cwd() / ".env"
        if not env_path.exists():
            # Try project root (walk up to find .env)
            for parent in Path.cwd().parents:
                candidate = parent / ".env"
                if candidate.exists():
                    env_path = candidate
                    break
        if env_path.exists():
            load_dotenv(env_path)
    except ImportError:
        pass  # dotenv not installed, rely on environment


# Load .env on import
load_env()


def check_openpyxl_available() -> bool:
    """Check if openpyxl is available."""
    try:
        import openpyxl
        return True
    except ImportError:
        return False


def check_gemini_available() -> bool:
    """Check if Gemini SDK is available."""
    try:
        from google import genai
        return True
    except ImportError:
        return False


def get_gemini_client():
    """Initialize Gemini client."""
    from google import genai

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY environment variable not set")

    return genai.Client(api_key=api_key)


def sanitize_filename(name: str) -> str:
    """Convert sheet name to safe filename."""
    # Replace spaces and special chars
    safe = re.sub(r'[^\w\-]', '_', name)
    # Remove consecutive underscores
    safe = re.sub(r'_+', '_', safe)
    return safe.lower().strip('_')


def analyze_sheet_structure(ws) -> dict:
    """Analyze a worksheet to understand its structure."""
    from openpyxl.utils import get_column_letter

    # Get dimensions
    min_row = ws.min_row or 1
    max_row = ws.max_row or 1
    min_col = ws.min_column or 1
    max_col = ws.max_column or 1

    # Sample data for structure analysis
    sample_rows = []
    for row_idx in range(min_row, min(min_row + 10, max_row + 1)):
        row_data = []
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append({
                "value": str(cell.value) if cell.value is not None else "",
                "is_merged": False,  # Would need merge check
                "has_formula": str(cell.value).startswith("=") if cell.value else False
            })
        sample_rows.append(row_data)

    # Detect if first row looks like headers
    first_row = sample_rows[0] if sample_rows else []
    has_headers = all(
        cell["value"] and not cell["value"].replace(".", "").replace("-", "").isdigit()
        for cell in first_row[:min(5, len(first_row))]
        if cell["value"]
    )

    # Check for merged cells
    merged_ranges = list(ws.merged_cells.ranges) if hasattr(ws, 'merged_cells') else []

    # Detect multiple tables (gaps in data)
    tables_detected = 1  # Basic detection

    return {
        "dimensions": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
        "rows": max_row - min_row + 1,
        "columns": max_col - min_col + 1,
        "has_headers": has_headers,
        "merged_cell_count": len(merged_ranges),
        "sample_data": sample_rows[:5],
        "tables_detected": tables_detected,
        "complexity": "high" if len(merged_ranges) > 5 or max_col > 20 else "standard"
    }


def extract_sheet_to_csv(ws, output_path: Path, has_headers: bool = True) -> dict:
    """Extract worksheet to CSV file."""
    rows_written = 0

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        for row in ws.iter_rows():
            row_values = []
            for cell in row:
                value = cell.value
                if value is None:
                    row_values.append("")
                elif isinstance(value, (int, float)):
                    row_values.append(value)
                else:
                    # Clean string values
                    row_values.append(str(value).strip())

            # Skip completely empty rows
            if any(v != "" for v in row_values):
                writer.writerow(row_values)
                rows_written += 1

    return {
        "path": str(output_path),
        "rows": rows_written,
        "has_headers": has_headers
    }


def sheet_to_markdown(ws, sheet_name: str, structure: dict) -> str:
    """Convert worksheet to markdown representation."""
    lines = [f"# {sheet_name}\n"]

    # Add structure summary
    lines.append("## Structure\n")
    lines.append(f"- **Dimensions:** {structure['dimensions']}")
    lines.append(f"- **Rows:** {structure['rows']}")
    lines.append(f"- **Columns:** {structure['columns']}")
    lines.append(f"- **Has Headers:** {'Yes' if structure['has_headers'] else 'No'}")
    lines.append(f"- **Complexity:** {structure['complexity']}")
    if structure['merged_cell_count'] > 0:
        lines.append(f"- **Merged Cells:** {structure['merged_cell_count']}")
    lines.append("")

    # Add data preview as markdown table
    lines.append("## Data Preview\n")

    # Get first 20 rows for preview
    preview_rows = []
    for idx, row in enumerate(ws.iter_rows(max_row=20)):
        row_values = []
        for cell in row:
            value = cell.value
            if value is None:
                row_values.append("")
            else:
                # Truncate long values
                str_val = str(value)[:50]
                if len(str(value)) > 50:
                    str_val += "..."
                # Escape pipe characters for markdown tables
                str_val = str_val.replace("|", "\\|")
                row_values.append(str_val)
        preview_rows.append(row_values)

    if preview_rows:
        # Build markdown table
        col_count = max(len(row) for row in preview_rows)

        # Header row
        if structure['has_headers'] and preview_rows:
            headers = preview_rows[0]
            lines.append("| " + " | ".join(headers + [""] * (col_count - len(headers))) + " |")
            lines.append("| " + " | ".join(["---"] * col_count) + " |")
            data_rows = preview_rows[1:]
        else:
            lines.append("| " + " | ".join([f"Col{i+1}" for i in range(col_count)]) + " |")
            lines.append("| " + " | ".join(["---"] * col_count) + " |")
            data_rows = preview_rows

        for row in data_rows:
            padded = row + [""] * (col_count - len(row))
            lines.append("| " + " | ".join(padded) + " |")

        if structure['rows'] > 20:
            lines.append(f"\n*... {structure['rows'] - 20} more rows*")

    lines.append("")
    return "\n".join(lines)


def analyze_with_gemini(client, workbook_info: dict, model: str = "gemini-2.0-flash") -> dict:
    """Use Gemini to analyze workbook structure and suggest extraction strategy."""

    prompt = f"""Analyze this Excel workbook structure and provide extraction guidance:

Workbook Info:
{json.dumps(workbook_info, indent=2)}

Please analyze and respond with JSON containing:
1. "relationships": Array of relationships between sheets (if any)
2. "extraction_notes": Object with sheet name keys, containing notes about how to best extract each sheet
3. "complexity_assessment": Overall complexity (simple/moderate/complex)
4. "recommended_approach": Brief recommendation for processing

Focus on:
- Are there lookup tables that relate to main data tables?
- Are there multi-level headers that need flattening?
- Are there embedded sub-tables within sheets?
- What columns should be split into separate CSVs?

Respond with valid JSON only, no markdown formatting."""

    try:
        response = client.models.generate_content(
            model=model,
            contents=[prompt]
        )

        # Try to parse as JSON
        text = response.text.strip()
        # Remove markdown code blocks if present
        if text.startswith("```"):
            text = re.sub(r'^```\w*\n?', '', text)
            text = re.sub(r'\n?```$', '', text)

        return json.loads(text)

    except Exception as e:
        return {
            "error": str(e),
            "relationships": [],
            "extraction_notes": {},
            "complexity_assessment": "unknown",
            "recommended_approach": "Standard extraction - Gemini analysis failed"
        }


def process_workbook(
    file_path: Path,
    output_dir: Path,
    sheets: list[str] | None = None,
    use_gemini: bool = True,
    model: str = "gemini-2.0-flash"
) -> dict:
    """Process entire Excel workbook."""
    from openpyxl import load_workbook

    output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbook
    wb = load_workbook(file_path, read_only=False, data_only=True)

    # Filter sheets if specified
    sheet_names = sheets if sheets else wb.sheetnames

    # Analyze all sheets
    workbook_info = {
        "file": file_path.name,
        "sheets": {}
    }

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        workbook_info["sheets"][sheet_name] = analyze_sheet_structure(ws)

    # Get Gemini analysis if available
    gemini_analysis = None
    if use_gemini and check_gemini_available():
        try:
            client = get_gemini_client()
            gemini_analysis = analyze_with_gemini(client, workbook_info, model)
        except Exception as e:
            gemini_analysis = {"error": str(e)}

    # Process each sheet
    files_created = []
    sheet_results = {}

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        safe_name = sanitize_filename(sheet_name)
        structure = workbook_info["sheets"][sheet_name]

        # Create markdown representation
        md_content = sheet_to_markdown(ws, sheet_name, structure)
        md_path = output_dir / f"sheet_{safe_name}.md"
        md_path.write_text(md_content)
        files_created.append(str(md_path))

        # Extract to CSV
        csv_path = output_dir / f"sheet_{safe_name}.csv"
        csv_info = extract_sheet_to_csv(ws, csv_path, structure['has_headers'])
        files_created.append(str(csv_path))

        sheet_results[sheet_name] = {
            "markdown": str(md_path),
            "csv": csv_info,
            "structure": structure
        }

    # Write index file
    index_content = [f"# {file_path.name}\n"]
    index_content.append("## Workbook Overview\n")
    index_content.append(f"- **Sheets:** {len(sheet_results)}")
    index_content.append(f"- **Source:** {file_path.name}\n")

    index_content.append("## Sheets\n")
    for sheet_name, info in sheet_results.items():
        safe_name = sanitize_filename(sheet_name)
        index_content.append(f"### {sheet_name}\n")
        index_content.append(f"- [Markdown](./sheet_{safe_name}.md)")
        index_content.append(f"- [CSV](./sheet_{safe_name}.csv)")
        index_content.append(f"- Rows: {info['structure']['rows']}, Columns: {info['structure']['columns']}")
        index_content.append(f"- Complexity: {info['structure']['complexity']}\n")

    if gemini_analysis and "recommended_approach" in gemini_analysis:
        index_content.append("## AI Analysis\n")
        index_content.append(f"**Recommendation:** {gemini_analysis.get('recommended_approach', 'N/A')}\n")

        if gemini_analysis.get("relationships"):
            index_content.append("### Relationships\n")
            for rel in gemini_analysis["relationships"]:
                index_content.append(f"- {rel}")
            index_content.append("")

    index_content.append("---\n*Processed by RFP Questions Document Ingestion*\n")

    index_path = output_dir / "index.md"
    index_path.write_text("\n".join(index_content))
    files_created.insert(0, str(index_path))

    # Write relationships JSON if we have Gemini analysis
    if gemini_analysis:
        rel_path = output_dir / "relationships.json"
        rel_path.write_text(json.dumps(gemini_analysis, indent=2))
        files_created.append(str(rel_path))

    wb.close()

    return {
        "success": True,
        "file": str(file_path),
        "output_dir": str(output_dir),
        "sheets_processed": len(sheet_results),
        "files_created": files_created,
        "gemini_analysis": gemini_analysis
    }


def main():
    parser = argparse.ArgumentParser(
        description="Process Excel files with intelligent extraction"
    )
    parser.add_argument("file_path", help="Path to Excel file (.xlsx)")
    parser.add_argument("--output-dir", "-o", required=True,
                        help="Output directory for extracted files")
    parser.add_argument("--sheets", "-s", default=None,
                        help="Comma-separated list of sheet names to process")
    parser.add_argument("--no-gemini", action="store_true",
                        help="Skip Gemini analysis")
    parser.add_argument("--model", "-m", default="gemini-2.5-flash",
                        help="Gemini model to use")
    parser.add_argument("--dry-run", action="store_true",
                        help="Analyze structure without processing")

    args = parser.parse_args()

    file_path = Path(args.file_path)
    output_dir = Path(args.output_dir)

    if not file_path.exists():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)

    if file_path.suffix.lower() not in (".xlsx", ".xls"):
        print(json.dumps({
            "error": f"Not an Excel file: {file_path.suffix}",
            "supported": [".xlsx", ".xls"]
        }))
        sys.exit(1)

    # Check openpyxl
    if not check_openpyxl_available():
        print(json.dumps({
            "error": "openpyxl package not installed",
            "fix": "pip install openpyxl"
        }))
        sys.exit(1)

    # Parse sheet list
    sheets = None
    if args.sheets:
        sheets = [s.strip() for s in args.sheets.split(",")]

    # Dry run - just analyze
    if args.dry_run:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)

        info = {
            "dry_run": True,
            "file": str(file_path),
            "sheets": {}
        }

        for sheet_name in wb.sheetnames:
            if sheets and sheet_name not in sheets:
                continue
            ws = wb[sheet_name]
            info["sheets"][sheet_name] = analyze_sheet_structure(ws)

        wb.close()
        print(json.dumps(info, indent=2))
        sys.exit(0)

    # Process workbook
    try:
        result = process_workbook(
            file_path,
            output_dir,
            sheets=sheets,
            use_gemini=not args.no_gemini,
            model=args.model
        )
        print(json.dumps(result, indent=2))

    except Exception as e:
        print(json.dumps({"error": f"Processing failed: {str(e)}"}))
        sys.exit(1)


if __name__ == "__main__":
    main()
